import io
from django.http import HttpResponse
from django.utils import timezone
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BRAND_NAME = "Agroveterinaria Planeta Animal"
BRAND_DARK = colors.HexColor("#144115")
BRAND_GREEN = colors.HexColor("#39701B")
BRAND_SOFT = colors.HexColor("#F4F8F1")
BRAND_TEXT = colors.HexColor("#526452")


def _pdf_footer(canvas, doc):
    page_width, _ = doc.pagesize
    canvas.saveState()
    canvas.setStrokeColor(BRAND_GREEN)
    canvas.setLineWidth(0.4)
    canvas.line(doc.leftMargin, 1.12 * cm, page_width - doc.rightMargin, 1.12 * cm)
    canvas.setFillColor(BRAND_TEXT)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawString(doc.leftMargin, 0.72 * cm, BRAND_NAME + " · Reporte generado por el sistema")
    canvas.drawRightString(
        page_width - doc.rightMargin,
        0.72 * cm,
        f"Página {canvas.getPageNumber()}",
    )
    canvas.restoreState()



# ─── VENTAS PDF ───────────────────────────────────────────────────────────────

def export_ventas_pdf(sales, fecha_desde=None, fecha_hasta=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('titulo', parent=styles['Normal'],
                                   fontSize=17, leading=21, fontName='Helvetica-Bold', textColor=BRAND_DARK, alignment=TA_CENTER)
    estilo_sub = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=9, leading=13, textColor=BRAND_TEXT, alignment=TA_CENTER)
    estilo_center = ParagraphStyle('center', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9)

    elementos = []

    elementos.append(Paragraph("Agroveterinaria Planeta Animal", estilo_titulo))
    elementos.append(Paragraph("Reporte de Ventas", estilo_sub))

    if fecha_desde and fecha_hasta:
        elementos.append(Paragraph(
            f"Periodo: {fecha_desde} - {fecha_hasta}", estilo_sub))
    else:
        elementos.append(Paragraph(
            f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", estilo_sub))

    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=BRAND_DARK))
    elementos.append(Spacer(1, 0.4*cm))

    # Resumen
    ventas_activas = [s for s in sales if not s.is_returned]
    total_general = sum(s.total for s in ventas_activas)
    resumen = [
        ["Total facturas", "Devoluciones", "Total recaudado"],
        [str(len(ventas_activas)), str(len(sales) - len(ventas_activas)),
         "$ {:,}".format(total_general).replace(",", ".")],
    ]
    tabla_resumen = Table(resumen, colWidths=[8*cm, 8*cm, 8*cm])
    tabla_resumen.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 0.4*cm))

    # Detalle
    datos = [["Factura", "Fecha", "Cliente", "ID Cliente", "Cajero", "Total", "Estado"]]
    for sale in sales:
        datos.append([
            sale.invoice_number,
            timezone.localtime(sale.created_at).strftime('%d/%m/%Y %H:%M'),
            sale.customer_name,
            sale.customer_id,
            sale.cashier.full_name,
            "$ {:,}".format(sale.total).replace(",", "."),
            "Devuelto" if sale.is_returned else "Pagado",
        ])

    tabla = Table(datos, colWidths=[3*cm, 3.8*cm, 4.5*cm, 3*cm, 4.5*cm, 3*cm, 2.5*cm])
    tabla.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BRAND_SOFT]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)

    doc.build(elementos, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
    return response


# ─── VENTAS EXCEL ─────────────────────────────────────────────────────────────

def export_ventas_excel(sales, fecha_desde=None, fecha_hasta=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    color_header = "144115"
    color_alt = "F4F8F1"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=color_header)
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # Titulo
    ws.merge_cells("A1:G1")
    ws["A1"] = "Agroveterinaria Planeta Animal - Reporte de Ventas"
    ws["A1"].font = Font(bold=True, size=14, color=color_header)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    periodo = f"Periodo: {fecha_desde} - {fecha_hasta}" if fecha_desde and fecha_hasta else \
        f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"] = periodo
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="888888", size=9)

    # Resumen
    ws["A4"] = "Total facturas"
    ws["B4"] = "Devoluciones"
    ws["C4"] = "Total recaudado"
    ventas_activas = [s for s in sales if not s.is_returned]
    ws["A5"] = len(ventas_activas)
    ws["B5"] = len(sales) - len(ventas_activas)
    ws["C5"] = sum(s.total for s in ventas_activas)

    for col in ["A", "B", "C"]:
        ws[f"{col}4"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col}4"].fill = PatternFill("solid", fgColor=color_header)
        ws[f"{col}4"].alignment = header_align
        ws[f"{col}5"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"].font = Font(bold=True)

    # Headers
    headers = ["Factura", "Fecha", "Cliente", "ID Cliente", "Cajero", "Total", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Datos
    for row_idx, sale in enumerate(sales, 8):
        fill = PatternFill("solid", fgColor=color_alt) if row_idx % 2 == 0 else None
        valores = [
            sale.invoice_number,
            timezone.localtime(sale.created_at).strftime('%d/%m/%Y %H:%M'),
            sale.customer_name,
            sale.customer_id,
            sale.cashier.full_name,
            sale.total,
            "Devuelto" if sale.is_returned else "Pagado",
        ]
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=valor)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill

    # Ancho columnas
    anchos = [15, 20, 25, 15, 25, 15, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.row_dimensions[7].height = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
    return response


# ─── INVENTARIO PDF ───────────────────────────────────────────────────────────

def export_inventario_pdf(products):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('titulo', parent=styles['Normal'],
                                   fontSize=17, leading=21, fontName='Helvetica-Bold', textColor=BRAND_DARK, alignment=TA_CENTER)
    estilo_sub = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=9, leading=13, textColor=BRAND_TEXT, alignment=TA_CENTER)

    elementos = []
    elementos.append(Paragraph("Agroveterinaria Planeta Animal", estilo_titulo))
    elementos.append(Paragraph("Reporte de Inventario", estilo_sub))
    elementos.append(Paragraph(
        f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}", estilo_sub))
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(HRFlowable(width="100%", thickness=1, color=BRAND_DARK))
    elementos.append(Spacer(1, 0.4*cm))

    datos = [["Codigo", "Producto", "Costo", "Precio Venta", "Stock", "Stock Min.", "Estado"]]
    for p in products:
        if p.stock == 0:
            estado = "Agotado"
        elif p.stock <= p.min_stock:
            estado = "Stock Bajo"
        else:
            estado = "Disponible"

        datos.append([
            p.code,
            p.name,
            "$ {:,}".format(p.cost_price).replace(",", "."),
            "$ {:,}".format(p.sell_price).replace(",", "."),
            str(p.stock),
            str(p.min_stock),
            estado,
        ])

    tabla = Table(datos, colWidths=[2.5*cm, 6*cm, 2.8*cm, 2.8*cm, 1.8*cm, 2.2*cm, 2.5*cm])
    tabla.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BRAND_SOFT]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)

    doc.build(elementos, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'
    return response


# ─── INVENTARIO EXCEL ─────────────────────────────────────────────────────────

def export_inventario_excel(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    color_header = "144115"
    color_alt = "F4F8F1"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=color_header)
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "Agroveterinaria Planeta Animal - Reporte de Inventario"
    ws["A1"].font = Font(bold=True, size=14, color=color_header)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="888888", size=9)

    headers = ["Codigo", "Producto", "Costo", "Precio Venta", "Stock", "Stock Min.", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx, p in enumerate(products, 5):
        if p.stock == 0:
            estado = "Agotado"
            color_estado = "dc2626"
        elif p.stock <= p.min_stock:
            estado = "Stock Bajo"
            color_estado = "d97706"
        else:
            estado = "Disponible"
            color_estado = "16a34a"

        fill = PatternFill("solid", fgColor=color_alt) if row_idx % 2 == 0 else None
        valores = [p.code, p.name, p.cost_price, p.sell_price, p.stock, p.min_stock, estado]

        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=valor)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill

        ws.cell(row=row_idx, column=7).font = Font(bold=True, color=color_estado)

    anchos = [12, 30, 15, 15, 10, 12, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws.row_dimensions[4].height = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'
    return response


def export_rentabilidad_excel(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rentabilidad"

    color_header = "144115"
    color_alt = "F4F8F1"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=color_header)
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "Agroveterinaria Planeta Animal - Reporte de Rentabilidad"
    ws["A1"].font = Font(bold=True, size=14, color=color_header)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="888888", size=9)

    total_cost = sum(p.cost_price * p.stock for p in products)
    total_sale_value = sum(p.sell_price * p.stock for p in products)
    total_profit = total_sale_value - total_cost
    overall_margin = (total_profit / total_sale_value * 100) if total_sale_value else 0

    summary = [
        ("Costo en stock", total_cost),
        ("Valor de venta", total_sale_value),
        ("Utilidad potencial", total_profit),
        ("Margen general", f"{overall_margin:.1f}%"),
    ]
    for col, (label, value) in enumerate(summary, 1):
        header_cell = ws.cell(row=4, column=col, value=label)
        value_cell = ws.cell(row=5, column=col, value=value)
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = header_fill
        header_cell.alignment = header_align
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.font = Font(bold=True)

    headers = [
        "Codigo",
        "Producto",
        "Costo",
        "Precio venta",
        "Utilidad und.",
        "Margen",
        "Stock",
        "Utilidad stock",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx, product in enumerate(products, 8):
        unit_profit = product.sell_price - product.cost_price
        margin = (unit_profit / product.sell_price * 100) if product.sell_price else 0
        stock_profit = unit_profit * product.stock
        fill = PatternFill("solid", fgColor=color_alt) if row_idx % 2 == 0 else None
        values = [
            product.code,
            product.name,
            product.cost_price,
            product.sell_price,
            unit_profit,
            f"{margin:.1f}%",
            product.stock,
            stock_profit,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill

        profit_color = "dc2626" if unit_profit < 0 else "16a34a"
        margin_color = "dc2626" if margin < 0 else "d97706" if margin < 20 else "16a34a"
        ws.cell(row=row_idx, column=5).font = Font(bold=True, color=profit_color)
        ws.cell(row=row_idx, column=6).font = Font(bold=True, color=margin_color)
        ws.cell(row=row_idx, column=8).font = Font(bold=True, color=profit_color)

    widths = [12, 30, 15, 15, 15, 12, 10, 16]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.row_dimensions[7].height = 25
    ws.freeze_panes = "A8"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_rentabilidad.xlsx"'
    return response


def export_rentabilidad_pdf(products):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('titulo_rentabilidad', parent=styles['Normal'],
                                 fontSize=17, leading=21, fontName='Helvetica-Bold', textColor=BRAND_DARK, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('sub_rentabilidad', parent=styles['Normal'],
                                    fontSize=9, leading=13, textColor=BRAND_TEXT, alignment=TA_CENTER)

    total_cost = sum(p.cost_price * p.stock for p in products)
    total_sale_value = sum(p.sell_price * p.stock for p in products)
    total_profit = total_sale_value - total_cost
    overall_margin = (total_profit / total_sale_value * 100) if total_sale_value else 0

    elements = [
        Paragraph("Agroveterinaria Planeta Animal", title_style),
        Paragraph("Reporte de Rentabilidad", subtitle_style),
        Paragraph(
            f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}",
            subtitle_style,
        ),
        Spacer(1, 0.4*cm),
        HRFlowable(width="100%", thickness=1, color=BRAND_DARK),
        Spacer(1, 0.4*cm),
    ]

    summary = [
        ["Costo en stock", "Valor de venta", "Utilidad potencial", "Margen general"],
        [
            "$ {:,}".format(total_cost).replace(",", "."),
            "$ {:,}".format(total_sale_value).replace(",", "."),
            "$ {:,}".format(total_profit).replace(",", "."),
            f"{overall_margin:.1f}%",
        ],
    ]
    summary_table = Table(summary, colWidths=[6.5*cm, 6.5*cm, 6.5*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.4*cm))

    data = [["Codigo", "Producto", "Costo", "Venta", "Utilidad und.", "Margen", "Stock", "Utilidad stock"]]
    for product in products:
        unit_profit = product.sell_price - product.cost_price
        margin = (unit_profit / product.sell_price * 100) if product.sell_price else 0
        stock_profit = unit_profit * product.stock
        data.append([
            product.code,
            product.name,
            "$ {:,}".format(product.cost_price).replace(",", "."),
            "$ {:,}".format(product.sell_price).replace(",", "."),
            "$ {:,}".format(unit_profit).replace(",", "."),
            f"{margin:.1f}%",
            str(product.stock),
            "$ {:,}".format(stock_profit).replace(",", "."),
        ])

    table = Table(data, colWidths=[2.5*cm, 5.5*cm, 2.7*cm, 2.7*cm, 3*cm, 2.1*cm, 1.7*cm, 3.2*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BRAND_SOFT]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements, onFirstPage=_pdf_footer, onLaterPages=_pdf_footer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_rentabilidad.pdf"'
    return response
    
